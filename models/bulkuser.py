from playwright.sync_api import expect
from models.user import UserListPage
import openpyxl

class BulkUserCreatePage:
    def __init__(self, page, readBulkUserID):
        self.page = page
        self.url = "https://stage-dms.robi.com.bd/#/user/"
        self.id = readBulkUserID
        self.action = "Create"

    def navigateToUserListPage(self):
        UserListPage.navigateToUserList(self)

    def navigateToBulkUserCreate(self):
        expect(self.page.get_by_role("link", name=f"Bulk User {self.action}")).to_be_visible()
        self.page.get_by_role("link", name=f"Bulk User {self.action}").click()
        self.page.get_by_label("Document(Only .xls, .xlsx) *").click()

    def selectMissingColumnsFile(self):
        self.page.get_by_label("Document(Only .xls, .xlsx) *").set_input_files(r"files\Bulk_User_"+self.action+"_Missing_Columns.xlsx")
        self.page.get_by_role("button", name="Upload").click()
        expect(self.page.get_by_text("does not")).to_be_visible()

    def modifyValidBulkUsers(self):
        from openpyxl import load_workbook

        xlsx = load_workbook(filename=r"files\Bulk_User_Create_Valid.xlsx")
        sheet = xlsx.active
        for i in range(2,7):
            sheet[f"A{i}"] = "BulkUser"+str(int(self.id)+1)
            sheet[f"C{i}"] = '0' + str(1800000000 + int(self.id) + 1)
            self.id = int(self.id)+1

        xlsx.save(filename=r"files\Bulk_User_Create_Valid.xlsx")

    def uploadValidFile(self):
        self.modifyValidBulkUsers()
        self.page.get_by_label("Document(Only .xls, .xlsx) *").click()
        self.page.get_by_label("Document(Only .xls, .xlsx) *").set_input_files(r"files\Bulk_User_"+self.action+"_Valid.xlsx")
        self.page.get_by_role("button", name="Upload").click()

    def confirmCreatedBulkUsers(self):
        expect(self.page.get_by_text("Uploaded Successfully...")).to_be_visible()
        self.page.get_by_text("Uploaded Successfully...").click()
        expect(self.page.locator("tr:nth-child(2) td:first-child")).to_contain_text("BulkUser"+str(self.id))
        for i in range(1,5):
            expect(self.page.locator(f"tr:nth-child({2+i}) td:first-child")).to_contain_text("BulkUser"+str(int(self.id)-i))

        def writeBulkUserID():
            with open(r"files\bulkuser_id.txt",'r') as f:
                st = int(f.read())
            with open(r"files\bulkuser_id.txt",'w') as f:
                nst = st + 5
                f.write(str(nst))
        writeBulkUserID()

class BulkUserUpdatePage(BulkUserCreatePage):
    def __init__(self, page):
        self.page = page
        self.url = "https://stage-dms.robi.com.bd/#/user/"
        self.action = "Update"

    def navigateToBulkUserUpdate(self):
        return super().navigateToBulkUserCreate()

    def duplicateValidFile(self):
        import shutil

        shutil.copy2(r"files\Bulk_User_Create_Valid.xlsx", r"files\Bulk_User_Update_Valid.xlsx")

    def deleteColumnFromValidFile(self):
        from openpyxl import load_workbook

        self.duplicateValidFile()

        xlsx = load_workbook(filename=r"files\Bulk_User_Update_Valid.xlsx")
        sheet = xlsx.active
        
        sheet.delete_cols(9)

        xlsx.save(filename=r"files\Bulk_User_Update_Missing_Columns.xlsx")

    def createMissingColumnsFile(self):
        self.deleteColumnFromValidFile()

    def modifyValidBulkUsers(self):
        from openpyxl import load_workbook

        xlsx = load_workbook(filename=r"files\Bulk_User_Update_Valid.xlsx")
        sheet = xlsx.active
        for i in range(2,7):
            sheet[f"B{i}"] = sheet[f"A{i}"].value + "_edited"

        xlsx.save(filename=r"files\Bulk_User_Update_Valid.xlsx")

    def confirmUpdatedBulkUsers(self):
        expect(self.page.get_by_text("Uploaded Successfully...")).to_be_visible()
        self.page.get_by_text("Uploaded Successfully...").click()